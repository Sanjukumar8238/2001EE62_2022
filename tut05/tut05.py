import pandas as pd
import os
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl.styles.borders import Border, Side
from datetime import datetime
start_time = datetime.now()

#Help https://youtu.be/N6PBd4XdnEw
def octant_range_names(mod=5000):
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    def find_octant(a,b,c):                                         # Function to find the octant 
        if(a>0 and b>0 and c>0):    
            return 1
        elif(a>0 and b>0 and c<0):
            return -1
        elif(a<0 and b>0 and c>0):
            return 2
        elif(a<0 and b>0 and c<0):
            return -2
        elif(a<0 and b<0 and c>0):
            return 3
        elif(a<0 and b<0 and c<0):
            return -3
        elif(a>0 and b<0 and c>0):
            return 4
        elif(a>0 and b<0 and c<0):
            return -4

    try:
        df=pd.read_excel('octant_input.xlsx')       # Reading input file and storing in dataframe 'df'
    except:
        print("File opening error")
        exit()

    n=len(df['U'])                                                  # Finding number of values

    try:
        u_avg=df['U'].mean()                                        # Finding average of u,v and w                 
        v_avg=df['V'].mean()
        w_avg=df['W'].mean()
    except:
        print("Error in values of points")
        exit()

    l1=[u_avg]                                                      # Storing values of average of u,v, and w in lists
    l2=[v_avg]
    l3=[w_avg]
    for i in range(1,n):                                            # Appending empty string in all th remaining positions of list
        l1.append(" ")
        l2.append(" ")
        l3.append(" ")
    df['U Avg']=l1                                                  # Adding the lists as column in dataframe
    df['V Avg']=l2
    df['W Avg']=l3

    df[r"U'=U - U Avg"]=df['U']-u_avg                               # Subtracting mean of u,v and w from original values and creating a column of it
    df[r"V'=V - V Avg"]=df['V']-v_avg
    df[r"W'=W - W Avg"]=df['W']-w_avg

    octant=[]                                                       # List to store octants

    for i in df.index:                                              # Finding octant and storing them in a list named 'Octant'
        octant.append(find_octant(df[r"U'=U - U Avg"][i],df[r"V'=V - V Avg"][i],df[r"W'=W - W Avg"][i]))

    df['Octant']=octant                                             # Creating a column for storing corresponding octants in dataframe

    try:
        df.to_excel('output_octant_transition_identify.xlsx',index=False)                          # Saving the dataframe in file
    except:
        print("Error in writing to output file")
        exit()

    try:
        wb=load_workbook('output_octant_transition_identify.xlsx')                                 # Loading the file in workbook
    except:
        print("Error in loading output file")
        exit()

    ws=wb.active
    ws['L3']='User Input'                                           # Putting the string 'User Input' at its specified place

    matrix=[]                                                       # 2-d matrix for storing octants within ranges
    count=[0]*9                                                     # Creating a list for storing elements of 9 columns

    count[0]='Octant ID'                                            # Storing header list in 'count' list



    for i in range(0,4):
        count[2*i+1]=(i+1)
        count[2*(i+1)]=-(i+1)
    matrix.append(count)                                            # Appending header list in matrix
    for i in range(13,22):                                          # Writing header list in worksheet
        ws.cell(row=1,column=i).value=count[i-13]
    count=[0]*9                                                     # Resetting values in list 'count'

    for i in octant:                                                # Finding total count of values in different octants
        if(i==1):
            count[1]=count[1]+1
        elif(i==-1):
            count[2]=count[2]+1
        elif(i==2):
            count[3]=count[3]+1
        elif(i==-2):
            count[4]=count[4]+1
        elif(i==3):
            count[5]=count[5]+1
        elif(i==-3):
            count[6]=count[6]+1
        elif(i==4):
            count[7]=count[7]+1
        elif(i==-4):
            count[8]=count[8]+1

    count[0]='Overall Count'                                        # Creating overall count row
    matrix.append(count)                                           
    for i in range(13,22):                                          # Writing overall count in worksheet
        ws.cell(row=2,column=i).value=count[i-13]
    ws.cell(row=3,column=13).value='Mod '+str(mod)                  # Writing mod value at specified cell
                            


    n=len(octant)                                                   # Finding the number of points given in the input
    count=[0]*9                                                     # Resetting the values in the list 'count'
    k=0                                                             # Variable to keep track of the index of data we are on
    j=4                                                             # Variable to keep track of row in worksheet
    for i in octant:                                                # Counting number of values in different octants in mod range
        if(i==1):
            count[1]=count[1]+1
        elif(i==-1):
            count[2]=count[2]+1
        elif(i==2):
            count[3]=count[3]+1
        elif(i==-2):
            count[4]=count[4]+1
        elif(i==3):
            count[5]=count[5]+1
        elif(i==-3):
            count[6]=count[6]+1
        elif(i==4):
            count[7]=count[7]+1
        elif(i==-4):
            count[8]=count[8]+1
        k=k+1                                                       # Incrementing the index tracking variable
        if(k%mod==1):                                               # Processing the mod values in the range and storing them in the list 'count'
            count[0]=str(k-1)+'-'                       
        elif(k%mod==0 or k==n):
            count[0]=count[0]+str(k-1)                              # Here count[0]-> represents the range and further elements of count represents the count in different octants
            for i in range(13,22):                                  # Writing the mod count of octant in worksheet
                ws.cell(row=j,column=i).value=count[i-13]
            j=j+1                                                   # Incrementing row
            matrix.append(count)
            count=[0]*9                                             # Resetting count of values in different octants

                                            

    try:
        wb.save('output_octant_transition_identify.xlsx')                                          # Saving the file
    except:
        print("Error in saving the output file")


from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000 
octant_range_names(mod)



#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
